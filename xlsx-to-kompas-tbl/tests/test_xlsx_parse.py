from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
FIXTURE_PATH = PROJECT_ROOT / "fixtures" / "table_M2.xlsx"
EXPECTED_PATH = PROJECT_ROOT / "fixtures" / "table_M2_expected_matrix.json"
LAYOUT_CONFIG_PATH = PROJECT_ROOT / "config" / "table_layout.ini"


def read_matrix(path: Path) -> tuple[list[list[str]], int, int]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook.worksheets[0]

    max_row = worksheet.max_row
    max_col = worksheet.max_column
    matrix: list[list[str]] = []

    for row in worksheet.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True
    ):
        matrix.append(["" if value is None else str(value) for value in row])

    workbook.close()
    return matrix, max_row, max_col


def read_layout_config(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith(("#", ";")):
            continue
        key, sep, value = line.partition("=")
        if sep:
            values[key.strip().lower()] = value.strip()
    return values


def test_fixture_dimensions() -> None:
    matrix, rows, cols = read_matrix(FIXTURE_PATH)

    assert rows == 8, "Ожидалось 8 строк (2 заголовка + 6 данных)."
    assert cols == 13, "Ожидалось 13 столбцов."
    assert len(matrix) == 8
    assert all(len(row) == 13 for row in matrix)


def test_control_values() -> None:
    matrix, _, _ = read_matrix(FIXTURE_PATH)

    assert matrix[0][0] == "\u0429\u0410\u041E"  # A1: ЩАО
    assert matrix[1][1] == "M2.2"   # B2
    assert matrix[2][0] == "1,80"   # A3
    assert matrix[6][10] == "321"   # K7
    assert matrix[7][0] == "<0,2"   # A8


def test_expected_matrix_snapshot() -> None:
    matrix, rows, cols = read_matrix(FIXTURE_PATH)
    expected = json.loads(EXPECTED_PATH.read_text(encoding="utf-8"))

    assert rows == expected["rows"]
    assert cols == expected["cols"]
    assert matrix == expected["matrix"]


def test_layout_config_defaults() -> None:
    assert LAYOUT_CONFIG_PATH.exists(), f"Не найден конфиг: {LAYOUT_CONFIG_PATH}"
    config = read_layout_config(LAYOUT_CONFIG_PATH)

    assert config.get("mode") in {"cell", "table"}
    assert float(config["cell_width_mm"]) > 0
    assert float(config["cell_height_mm"]) > 0
